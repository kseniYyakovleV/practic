#django
from django.shortcuts import render
from django.shortcuts import HttpResponse
from django.http import FileResponse
from django.core.files.storage import FileSystemStorage
from db.models import Element, User, Addings, Count_changings
from rest_framework.response import Response
from rest_framework import generics
#os
import os
#excel
import openpyxl
from openpyxl.drawing.image import Image
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment
#date
import datetime
from dateutil.relativedelta import relativedelta
import pytz
# Create your views here.



class get_valid_for_excel_objects(generics.ListAPIView):
    #Получить корректные для запроса записи (нужные поля заполнены)
    def get(self, request):
        return Response(Element.get_valid_for_excel_objects_data())
    

class get_not_valid_for_excel_objects(generics.ListAPIView):
    #Получить некорректные для запроса записи (не все нужные поля заполнены)
    def get(self, request):
        return Response(Element.get_not_valid_for_excel_objects_data())
    


class get_one_item_data(generics.ListAPIView):
    #Получить все данные одной записи
    def get(self, request):
        number = request.GET["number"]
        object = Element.objects.get(number=number)
        return Response(object.get_all_data_of_one_element())
    


class get_all_items_data(generics.ListAPIView):
    #Получить все данные всех записей
    def get(self, request):
        return Response(Element.get_all_data_of_all_elements())


def show_image(request):
    #Отправить изображение по полю "image" запроса GET
    image = request.GET["image"]
    response = FileResponse(open(os.path.abspath("db/files/images")+"/"+image, "rb"))
    return response

def empty_number():
        images = os.listdir(os.path.abspath(r"db/files/images"))
        number = 1
        for number in range(1, len(images)+1):
            if "image"+str(number)+".png" not in images and "image"+str(number)+".jpeg" not in images:
                return number
        return number+1


def add_data(request):
    response = HttpResponse("")
    if request.method == "POST":
        end = ""
        user = User.objects.filter(login=request.POST["login"], password=request.POST["password"])
        if user.exists():
            file = request.FILES["image"]
            response["status"] = "OK"
            new_element = Element(number=Element.get_free_number(), manufacturer=request.POST["manufacturer"],
                equipment=request.POST["equipment"], analog=request.POST["analog"], type=request.POST["type"],
                name=request.POST["name"], SAP_number=request.POST["SAP_number"], serial_number=request.POST["serial_number"],
                storage=request.POST["storage"], unit=request.POST["unit"],delivery_time=request.POST["delivery_time"],
                using_time=request.POST["using_time"], downtime=request.POST["downtime"], 
                delivery_time_and_coefficient=request.POST["delivery_time_and_coefficient"], 
                using_in_workshop=request.POST["using_in_workshop"], price_in_rubles=request.POST["price_in_rubles"],
                count=request.POST["count"], ordered=request.POST["ordered"], min_fact=request.POST["min_fact"],
                max_fact=request.POST["max_fact"], image=None)
            
            if file.name.endswith(".jpeg"):
                end = ".jpeg"
            elif file.name.endswith(".png"):
                end = ".png"
            elif file.name!="":
                response["status"] = "Wrong type of image"
        else:
            response["status"] = "Wrong login or password"
        if end!='':
            storage = FileSystemStorage(os.path.abspath(os.path.normpath(r"db/files/images")))
            name = "image"+str(empty_number())+end
            filename = storage.save(name, file)
            new_element.image = filename
    else:
        response["status"] = "Must be 'POST' method"
    if response["status"] == "OK":
        new_element.save()
        Addings.objects.create(added_element=new_element, user=user[0], datetime=datetime.datetime.now(pytz.timezone("Europe/Moscow")))
    return response


def get_excel(request):
    #Сгенерировать отчет и отправить его, как файл .xlsx
    double = Side(border_style = "medium", color = "000000")
    thin = Side(border_style = "thin", color = "000000")
    double_border = Border(left = double, right = double, top = double, bottom = double)
    thin_border = Border(left = thin, right = thin, top = thin, bottom = thin)
    doc=openpyxl.load_workbook(os.path.abspath("db/files/spreadsheets/template.xlsx"))
    sheets = doc.get_sheet_names()
    sheet = doc[sheets[0]]
    date = datetime.date.today()

    n = 1
    sheet["H5"]="Дата документа/Document date: "+date.strftime("%d.%m.%Y")
    items = Element.get_valid_for_excel_objects_data()
    for i in items.values():
        title = i["name"]
        image = i["image"]
        brand = i["manufacturer"]
        analog = i["analog"]
        unit = i["unit"]
        count = i["count"]
        ordered = i["ordered"]
        min = i["min_fact"]
        price = i["price_in_rubles"]
        if count+ordered<min:
            sheet["A"+str(n+10)] = str(n)
            sheet["B"+str(n+10)] = title
            sheet["B"+str(n+10)].border = thin_border
            

            image = Image(os.path.abspath("db/files/images")+"/"+image)
            image.height = 150
            image.width = 150
            sheet.add_image(image, "C"+str(n+10))
            sheet["C"+str(n+10)].border = thin_border
            sheet["D"+str(n+10)] = "Запасные части,  предназначается для службы главного механика, инициатором закупки является служба главного механика.\n\n"+title
            sheet["D"+str(n+10)].border = thin_border
            sheet["E"+str(n+10)] = brand
            sheet["E"+str(n+10)].border = thin_border
            sheet["F"+str(n+10)] = analog
            sheet["F"+str(n+10)].border = thin_border
            
            sheet.merge_cells("F"+str(n+10)+":G"+str(n+10))
            sheet["H"+str(n+10)]=unit
            sheet["H"+str(n+10)].border = thin_border
            sheet["I"+str(n+10)]=min-count
            sheet["I"+str(n+10)].border = thin_border
            deadline_date = date+relativedelta(months=2)
            sheet["J"+str(n+10)]=deadline_date.strftime("%d.%m.%Y")
            sheet["J"+str(n+10)].border = thin_border

            sheet["K"+str(n+10)]="M&U (SW)"
            sheet["K"+str(n+10)].border = thin_border
            sheet["L"+str(n+10)]="Spare Parts and Service / Запасные части и сервис"
            sheet["L"+str(n+10)].border = thin_border
            sheet["M"+str(n+10)]=str(str(price)+" Р.")
            sheet["M"+str(n+10)].border = thin_border
            sheet["N"+str(n+10)]="""Инициатор закупки Володи С.В.
Решение о закупки Володи С.В.
Согласование закупки Онуфриев С.Ю. 
Для быстрого ремонта оборудования в цехе сварки, в случае отказа в работе основного устройства, приобретается в рамках списка ключевых запасных частей.


Procurement initiator Volodya S.V.
The decision to purchase Volodya S.V.
Procurement approval Onufriev S.Yu.
For quick repair of equipment in the welding shop, in case of failure of the main device, it is purchased as part of the use of spare parts."""
            sheet["O"+str(n+10)]=str(price*(min-count-ordered))+"Р"
            sheet["A"+str(n+10)].border = double_border
            sheet["N"+str(n+10)].border = double_border
            """
            i.ordered = i.min-i.count
            i.save()
            """
            n+=1

    sheet.row_dimensions[n+10].height = 75
    sheet.row_dimensions[n+11].height = 42


    sheet["A"+str(n+10)].border = double_border
    sheet["A"+str(n+10)].alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    sheet.merge_cells("A"+str(n+10)+":A"+str(n+11))
    sheet["A"+str(n+10)] = "Исполнитель /Editor"
    
    sheet["B"+str(n+10)].border = double_border
    sheet["B"+str(n+10)] = "Ф.И.О. / Подпись    Full Name/ Signature"
    sheet["B"+str(n+11)].border = double_border

    sheet["C"+str(n+10)].border = double_border
    sheet["C"+str(n+10)] = 'Руководитель  отдела-клиента/Head of client department'
    sheet.merge_cells("C"+str(n+10)+":E"+str(n+11))

    
    sheet["F"+str(n+10)].border = double_border
    sheet.merge_cells("F"+str(n+10)+":G"+str(n+11))

    sheet["H"+str(n+10)].border = double_border
    sheet.merge_cells("H"+str(n+10)+":H"+str(n+11))
    sheet["H"+str(n+10)] = "Руководитель отдела закупок/ Head of Purchasing Section"

    sheet["I"+str(n+10)].border = double_border
    sheet["I"+str(n+10)] = 'Ф.И.О. / Подпись                Full Name/ Signature'
    sheet.merge_cells("I"+str(n+10)+":J"+str(n+10))
    sheet["I"+str(n+11)].border = double_border
    sheet.merge_cells("I"+str(n+11)+":J"+str(n+11))

    sheet["K"+str(n+10)].border = double_border
    sheet["K"+str(n+10)].alignment = Alignment(horizontal = "center", vertical = "center", wrapText = True)
    sheet["K"+str(n+10)] = 'Генеральный Директор /       General Director'
    sheet.merge_cells("K"+str(n+10)+":L"+str(n+11))


    sheet["M"+str(n+10)].border = double_border
    sheet["M"+str(n+10)] = 'Ф.И.О. / Подпись        Full Name/ Signature'
    sheet.merge_cells("M"+str(n+10)+":N"+str(n+10))
    sheet["M"+str(n+11)].border = double_border
    sheet.merge_cells("M"+str(n+11)+":N"+str(n+11))



    doc.save("request.xlsx")
    with open(os.path.abspath("request.xlsx"), "rb") as file:
        my_data = file.read()
        response = HttpResponse(my_data, headers = {
            "Content-Type": "application/vnd.ms-excel",
            "Content-Disposition": "attachment; filename = " + date.strftime("%d.%m.%Y") + ".xlsx"})
        return response

    
def change_data(request):
    response = HttpResponse()
    response["status"] = "OK"
    if request.method == "POST":
        try:
            user = User.objects.get(login = request.POST["login"], password = request.POST["password"])
            try:
                element = Element.objects.get(number = request.POST["number"])
                element.manufacturer=request.POST["manufacturer"]
                element.equipment=request.POST["equipment"]
                element.analog=request.POST["analog"]
                element.type=request.POST["type"]
                element.name=request.POST["name"]
                element.SAP_number=request.POST["SAP_number"]
                element.serial_number=request.POST["serial_number"]
                element.storage=request.POST["storage"]
                element.unit=request.POST["unit"]
                element.delivery_time=request.POST["delivery_time"]
                element.using_time=request.POST["using_time"]
                element.downtime=request.POST["downtime"]
                element.delivery_time_and_coefficient=request.POST["delivery_time_and_coefficient"]
                element.using_in_workshop=request.POST["using_in_workshop"]
                element.price_in_rubles=request.POST["price_in_rubles"]
                if element.count!=request.POST["count"]:
                    Count_changings.objects.create(changed_element = element, user=user, datetime=datetime.datetime.now(pytz.timezone("Europe/Moscow")), 
                                        count_changing = float(request.POST["count"])-element.count)
                    element.count=request.POST["count"]
                element.ordered=request.POST["ordered"]
                element.min_fact=request.POST["min_fact"]
                element.max_fact=request.POST["max_fact"]
                
                if "image" in request.FILES:
                    file = request.FILES["image"]
                    if file.name.endswith(".jpeg"):
                        end = ".jpeg"
                    elif file.name.endswith(".png"):
                        end = ".png"
                    elif file.name!="":
                        response["status"] = "Wrong type of image"
                    if end!='':
                        image = element.image
                        storage = FileSystemStorage(os.path.abspath(os.path.normpath(r"db/files/images")))
                        name = "image"+str(empty_number())+end
                        filename = storage.save(name, file)
                        element.image = filename
                        count_elements_with_this_image = Element.objects.filter(image=image).__len__()
                        if count_elements_with_this_image<=1:
                            os.remove(os.path.abspath("db/files/images")+"/"+image)
                elif "image" in request.POST:
                    element.image = request.POST["image"]
                
                element.save()
            except Element.DoesNotExist:
                response["status"] = "Wrong number"
        except User.DoesNotExist:
            response["status"] = "Wrong login or password"
    else:
        response["status"] = "Must be 'POST' method"
    return response



def change_count(request):
    response = HttpResponse()
    response["status"] = "Count is not ehough"
    if request.method == "POST":
        try:
            user = User.objects.get(login = request.POST["login"], password = request.POST["password"])
            try:
                element = Element.objects.get(number = request.POST["number"])
                new_count = float(request.POST["new_count"])
                old_count = element.count
                Count_changings.objects.create(changed_element=element, user=user, datetime=datetime.datetime.now(pytz.timezone("Europe/Moscow")), 
                                       count_changing = old_count-new_count)

                
                element.save()
                response["status"] = "OK"
            except Element.DoesNotExist:
                response["status"] = "Wrong number"
            except TypeError:
                response["status"] = "Wrong types of arguments"
        except User.DoesNotExist:
            response["status"] = "Wrong login or password"
    else:
        response["status"] = "Must be 'POST' method"
    return response


def get_all_images(request):
    images = os.listdir(os.path.abspath(r"db/files/images"))
    response = HttpResponse()
    for image in images[:-1]:
        response.write(image+"|")
    if len(images)>=1:
        response.write(images[-1])
    print(response.content)
    return response