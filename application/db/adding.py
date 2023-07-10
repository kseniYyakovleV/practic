import re
from openpyxl import load_workbook
from zipfile import ZipFile
import os
import shutil
from db.models import Element
from django.db.utils import IntegrityError
from distutils.dir_util import copy_tree, remove_tree

filename = os.path.abspath("db/files/spreadsheets/data.xlsx")
path_in_archive = "xl/drawings"
dir_with_images = "xl/media"
path_for_images_types = "xl/drawings/_rels/drawing1.xml.rels"
dir_for_images = os.path.abspath("db/files/images")

full_pattern = "\<xdr:twoCellAnchor\>.\</xdr:twoCellAnchor\>"

cell_rId = "drawing1.xml"
shutil.copy(filename, filename+".zip")
zip = ZipFile(filename, mode="r")
text = zip.read(path_in_archive+"/"+cell_rId).decode()

count = 0
for i in text.split("/xdr:twoCellAnchor"):
    if "rId" not in i:
        count+=1

def find_image_in_cell(text):
    rId_pattern = r"rId\d+"
    row_pattern = r"row\>\d+"
    row = (re.findall(pattern=rId_pattern, string=text)[0][3:])
    image = (re.findall(pattern=row_pattern, string=text)[0][4:])
    return (int(row), "image"+image)

file_with_images_types = zip.read(path_for_images_types).decode()
file_with_images_types = file_with_images_types.split("\" Type=\"http://schemas.openxmlformats.org/officeDocument/2006/relationships/image\" Target=\"../media/")

all_types_of_images = dict()
for i in file_with_images_types[1:-1]:
    print(i)
    type = i.split("\"/><Relationship Id=\"rId")
    all_types_of_images["image" + str(int(type[1])-1)] = type[0]
print(all_types_of_images)
all_images = dict()
for i in text.split("/xdr:twoCellAnchor"):
    if "rId" in i:
        row = find_image_in_cell(i)
        all_images[row[0]] = row[1]




os.makedirs(dir_for_images, exist_ok=True)
files = zip.namelist()
images = []
for file in files:
    if "image" in file:
        images.append(file)


zip.extractall(dir_for_images, images)

zip.close()

wb = load_workbook(filename)
sheet = wb[wb.sheetnames[0]]
row_index = 2
with open("errors.txt", "w") as error_file:
    with open("integrity_errors.txt", "w") as integrity_error_file:
        with open("key_errors.txt", "w") as key_error_file:
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == None:
                    continue
                try:
                        if row_index in all_images:
                            image = all_types_of_images[all_images[row_index]]
                        else:
                            image = None
                        Element.objects.create(
                            number = row[0].value,
                            manufacturer = row[1].value,
                            equipment = row[2].value,
                            type = row[3].value,
                            name = row[4].value,
                            SAP_number = row[5].value,
                            serial_number = row[6].value,
                            storage = row[7].value,
                            unit = row[8].value,
                            delivery_time = row[9].value,
                            using_time = row[11].value,
                            downtime = row[13].value,
                            delivery_time_and_coefficient = row[16].value,
                            using_in_workshop = row[18].value,
                            price_in_rubles = row[20].value,
                            count = row[26].value,
                            min_fact = row[27].value,
                            max_fact = row[28].value,
                            ordered = 0,
                            image = image)
                        
                except ValueError:
                    print(row_index, file=error_file)

                except IntegrityError:
                    print(row_index, file=integrity_error_file)
                except KeyError:
                    print(row_index, file=key_error_file)
                row_index+=1



copy_tree(os.path.abspath("db/files/images/xl/media"), os.path.abspath("db/files/images"))
remove_tree(os.path.abspath("db/files/images/xl"))
